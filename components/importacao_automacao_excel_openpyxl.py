"""IMPORTAÇÕES PARA AUTOMAÇÃO DE PLANILHAS COM EXCEL"""
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, NamedStyle
from openpyxl.utils.exceptions import InvalidFileException
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf

def carrega_excel(arquivoExcel):
    try:
        workbook = load_workbook(arquivoExcel)
        style_moeda = NamedStyle(name="estilo_moeda", number_format='_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-')
        if "estilo_moeda" not in workbook.named_styles:
            workbook.add_named_style(style_moeda)
        sheet = workbook.active
        return workbook, sheet, style_moeda
    except FileNotFoundError:
        print(f"Erro: O arquivo '{arquivoExcel}' não foi encontrado.")
        return None, None, None
    except InvalidFileException:
        print(f"Erro: O arquivo '{arquivoExcel}' está corrompido ou não é um arquivo Excel válido.")
        return None, None, None
    except Exception as e:
        print(f"Erro inesperado: {e}")
        return None, None, None

def converter_excel_para_pdf(caminho_excel, caminho_pdf, nome_planilha=None):
    if nome_planilha:
        # Carregar a planilha específica do arquivo Excel
        wb = load_workbook(caminho_excel)
        ws = wb[nome_planilha]
        # Converter a planilha em um DataFrame
        data = ws.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)
    else:
        # Carregar o primeiro arquivo excel para um DataFrame
        df = pd.read_excel(caminho_excel)
    
    # Gerar o PDF usando o matplotlib
    pdf = matplotlib.backends.backend_pdf.PdfPages(caminho_pdf)
    fig, ax = plt.subplots(figsize=(8.27, 11.69))  # A4 size in inches
    ax.axis('tight')
    ax.axis('off')

    table = ax.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1.2, 1.2)

    pdf.savefig(fig, bbox_inches='tight')
    plt.close(fig)
    pdf.close()